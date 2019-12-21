import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

wb = load_workbook('data/regions.xlsx')
ws = wb.active

df = pd.read_excel('data/all_shifts.xlsx')

df1 = df[['Sales Rep', 'Cost per', 'Units Sold']]
df1['Total'] = df1['Cost per'] * df1['Units Sold']


rows = dataframe_to_rows(df1, index=False)

for r_index , row in enumerate(rows, 1):
  for c_index, col in enumerate(row, 6):
    ws.cell(row=r_index, column=c_index, value=col)

wb.save('target/test.xlsx')