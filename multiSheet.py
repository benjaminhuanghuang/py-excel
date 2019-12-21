import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font

# Compbine data frames
df_1 = pd.read_excel('data/shifts.xlsx', sheet_name='Sheet')
df_2 = pd.read_excel('data/shifts.xlsx', sheet_name='Sheet1')
df_3 = pd.read_excel('data/shift_3.xlsx')

df_all = pd.concat([df_1, df_2, df_3], sort=False)

to_excel = df_all.to_excel('target/all_shifts.xlsx', index=None)

wb = load_workbook('target/all_shifts.xlsx')
ws = wb.active

total_col = ws['G1']
total_col.font = Font(bold=True)
total_col.value = 'Total'

e_col, f_col = ['E', 'F']
for row in range(2, 300):
  result_cell = 'G{}'.format(row)
  e_value = ws[e_col + str(row)].value
  f_value = ws[f_col + str(row)].value
  ws[result_cell]= e_value * f_value


wb.save('target/total.xlsx')