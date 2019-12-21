from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = Workbook()
# Get sheet
ws = wb.active
ws1 = wb.create_sheet('newSheet')
ws2 = wb.create_sheet('firstSheet', 0)

# ['firstSheet', 'MySheet', 'newSheet']
ws.title = 'MySheet'

print(wb.sheetnames)


# Load excel into wrokbook
wb2 = load_workbook('data/regions.xlsx')
new_sheet = wb2.create_sheet('NewSheet')
active_sheet = wb2.active

# Access cell
cell = active_sheet['A1']
print(cell.value)

# Set cell
active_sheet['A1'] = 0
wb2.save('data/regions.xlsx')

# Columns
cell_range = active_sheet['A1':'C1']    # get a list
cell_range = active_sheet['A': 'C']      # get col A and col C
col_c = active_sheet['C']


# Rows
row_range = active_sheet[1:5]


# iteration
for row in active_sheet.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
  for cell in row:
    print(cell)


