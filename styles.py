from openpyxl.workbook import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Color, Alignment, PatternFill, GradientFill, Border, Side
from openpyxl.styles import NamedStyle

wb = Workbook()
# Get sheet
ws = wb.active

for i in range (1, 20):
  ws.append(range(30))    # Add new rows



# Merge
ws.merge_cells("A1:B5")

ws.unmerge_cells("A1:B5")

ws.merge_cells(start_row=2, start_column=2, end_row=5, end_column=5)

# Color
cell = ws['B2']
cell.font = Font(color=colors.RED, size=20, italic=True)
cell.value = 'Merged cell'
cell.alignment = Alignment(horizontal='right', vertical='bottom')


cell.fill = GradientFill(stop=("000000", "FFFFFF"))

# Style
highlight = NamedStyle(name='highlight')
highlight.font = Font(bold=True)
border= Side(style='thick', color='000000')
highlight.border = Border(left=border, top=border)
highlight.file = PatternFill('solid', fgColor='FFFF00')

count = 0
for col in ws.iter_cols(min_col=8, min_row=1, max_col=30, max_row=30):
  col[count].style = highlight
  count = count + 1

wb.save('target/temp.xlsx')